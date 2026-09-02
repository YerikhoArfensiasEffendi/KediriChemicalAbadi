"""
PT KEDIRI CHEMICAL ABADI — EXECUTIVE BUSINESS & FINANCIAL CONSULTING TOOLKIT
===========================================================================
Modul analisis bisnis, valuasi finansial, dan pemodelan strategis sekelas CEO / Managing Director.
Standar Kualitas: ISO 9001:2015 & Sanitasi Metadata Eksekutif.
Author / Manager: Yerikho Arfensias Effendi
Company: PT Kediri Chemical Abadi
Direktur Utama: Yan Effendi
"""

import math
from typing import Dict, List, Optional, Tuple, Any
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# 1. VALUASI KORPORAT & COST OF CAPITAL (DCF, WACC, CAPM)
# ==============================================================================

def calculate_capm(risk_free_rate: float, beta: float, market_return: float) -> float:
    """
    Menghitung Cost of Equity menggunakan Capital Asset Pricing Model (CAPM).
    Formula: Ke = Rf + Beta * (Rm - Rf)
    Semua input dalam format desimal (contoh: 0.065 untuk 6.5%).
    """
    cost_of_equity = risk_free_rate + (beta * (market_return - risk_free_rate))
    return round(cost_of_equity, 6)


def calculate_wacc(
    cost_of_equity: float,
    cost_of_debt_pretax: float,
    tax_rate: float,
    equity_value: float,
    debt_value: float
) -> Dict[str, Any]:
    """
    Menghitung Weighted Average Cost of Capital (WACC).
    Formula: WACC = (E/V * Ke) + (D/V * Kd * (1 - T))
    """
    total_value = equity_value + debt_value
    if total_value <= 0:
        raise ValueError("Total value (Equity + Debt) harus lebih besar dari 0.")
    
    weight_equity = equity_value / total_value
    weight_debt = debt_value / total_value
    after_tax_cost_of_debt = cost_of_debt_pretax * (1.0 - tax_rate)
    
    wacc = (weight_equity * cost_of_equity) + (weight_debt * after_tax_cost_of_debt)
    
    return {
        "wacc": round(wacc, 6),
        "wacc_pct": round(wacc * 100, 2),
        "weight_equity": round(weight_equity, 4),
        "weight_debt": round(weight_debt, 4),
        "after_tax_cost_of_debt": round(after_tax_cost_of_debt, 6)
    }


def dcf_valuation(
    free_cash_flows: List[float],
    wacc: float,
    terminal_growth_rate: float,
    net_debt: float = 0.0,
    shares_outstanding: Optional[float] = None
) -> Dict[str, Any]:
    """
    Melakukan valuasi Discounted Cash Flow (DCF) standar Goldman Sachs / Morgan Stanley.
    Menggunakan Gordon Growth Model untuk Terminal Value.
    
    :param free_cash_flows: Proyeksi FCF (tahun 1 s/d N) dalam Rupiah/Mata Uang.
    :param wacc: Discount rate (format desimal, misal 0.11 untuk 11%).
    :param terminal_growth_rate: Tingkat pertumbuhan terminal jangka panjang (misal 0.025 untuk 2.5%).
    :param net_debt: Total Utang dikurangi Kas & Setara Kas.
    :param shares_outstanding: Jumlah lembar saham beredar (opsional).
    """
    if wacc <= terminal_growth_rate:
        raise ValueError("WACC harus lebih besar daripada Terminal Growth Rate.")
    
    pv_fcf_list = []
    for t, fcf in enumerate(free_cash_flows, start=1):
        pv = fcf / ((1.0 + wacc) ** t)
        pv_fcf_list.append(round(pv, 2))
    
    sum_pv_fcf = sum(pv_fcf_list)
    last_fcf = free_cash_flows[-1]
    
    # Terminal Value pada akhir periode proyeksi (Tahun N)
    terminal_fcf = last_fcf * (1.0 + terminal_growth_rate)
    terminal_value = terminal_fcf / (wacc - terminal_growth_rate)
    
    # Present Value dari Terminal Value
    n_years = len(free_cash_flows)
    pv_terminal_value = terminal_value / ((1.0 + wacc) ** n_years)
    
    # Enterprise Value (EV) & Equity Value
    enterprise_value = sum_pv_fcf + pv_terminal_value
    equity_value = enterprise_value - net_debt
    
    implied_share_price = None
    if shares_outstanding and shares_outstanding > 0:
        implied_share_price = equity_value / shares_outstanding
    
    return {
        "projection_years": n_years,
        "pv_discrete_cash_flows": pv_fcf_list,
        "sum_pv_cash_flows": round(sum_pv_fcf, 2),
        "terminal_value": round(terminal_value, 2),
        "pv_terminal_value": round(pv_terminal_value, 2),
        "pv_terminal_value_pct": round((pv_terminal_value / enterprise_value) * 100, 2) if enterprise_value > 0 else 0,
        "enterprise_value": round(enterprise_value, 2),
        "net_debt": round(net_debt, 2),
        "equity_value": round(equity_value, 2),
        "implied_share_price": round(implied_share_price, 2) if implied_share_price else None
    }


# ==============================================================================
# 2. UNIT ECONOMICS & SAAS/MANUFACTURING METRICS
# ==============================================================================

def calculate_unit_economics(
    cac: float,
    arpu_monthly: float,
    gross_margin_pct: float,
    churn_rate_monthly: float
) -> Dict[str, Any]:
    """
    Menghitung metrik fundamental Unit Economics bisnis.
    
    :param cac: Customer Acquisition Cost (Biaya akuisisi per pelanggan).
    :param arpu_monthly: Rata-rata pendapatan bulanan per pelanggan (ARPU).
    :param gross_margin_pct: Gross margin dalam desimal (misal 0.40 untuk 40%).
    :param churn_rate_monthly: Churn rate bulanan dalam desimal (misal 0.03 untuk 3%).
    """
    if churn_rate_monthly <= 0:
        churn_rate_monthly = 0.001  # Fallback to prevent divide by zero
        
    avg_customer_lifetime_months = 1.0 / churn_rate_monthly
    customer_lifetime_value = (arpu_monthly * gross_margin_pct) / churn_rate_monthly
    ltv_cac_ratio = customer_lifetime_value / cac if cac > 0 else 0
    
    monthly_gross_profit_per_user = arpu_monthly * gross_margin_pct
    cac_payback_months = cac / monthly_gross_profit_per_user if monthly_gross_profit_per_user > 0 else 0
    
    # Evaluasi Kesehatan Unit Economics Kelas CEO
    if ltv_cac_ratio >= 4.0:
        health_status = "EXCELLENT (Skalabilitas Sangat Tinggi, Siap Ekspansi Agresif)"
    elif ltv_cac_ratio >= 3.0:
        health_status = "HEALTHY (Standar Emas Industri, Unit Economics Sehat)"
    elif ltv_cac_ratio >= 1.5:
        health_status = "MODERATE (Perlu Optimasi Biaya Akuisisi / Retensi Pelanggan)"
    else:
        health_status = "CRITICAL (Membakar Uang Tidak Berkelanjutan, Segera Evaluasi Strategi)"
        
    return {
        "customer_lifetime_months": round(avg_customer_lifetime_months, 1),
        "customer_lifetime_value_ltv": round(customer_lifetime_value, 2),
        "cac": round(cac, 2),
        "ltv_cac_ratio": round(ltv_cac_ratio, 2),
        "cac_payback_months": round(cac_payback_months, 1),
        "monthly_gross_profit_per_user": round(monthly_gross_profit_per_user, 2),
        "health_status": health_status
    }


# ==============================================================================
# 3. WORKING CAPITAL & CASH CONVERSION CYCLE (CCC)
# ==============================================================================

def calculate_working_capital_cycle(
    cogs_annual: float,
    revenue_annual: float,
    average_inventory: float,
    average_receivables: float,
    average_payables: float,
    days_in_period: int = 365
) -> Dict[str, Any]:
    """
    Menghitung siklus modal kerja (Cash Conversion Cycle - CCC).
    Metrik vital untuk efisiensi likuiditas manufaktur dan distribusi.
    
    - DIO (Days Inventory Outstanding): Hari perputaran persediaan
    - DSO (Days Sales Outstanding): Hari penagihan piutang
    - DPO (Days Payable Outstanding): Hari pelunasan utang usaha
    - CCC = DIO + DSO - DPO
    """
    dio = (average_inventory / cogs_annual) * days_in_period if cogs_annual > 0 else 0
    dso = (average_receivables / revenue_annual) * days_in_period if revenue_annual > 0 else 0
    dpo = (average_payables / cogs_annual) * days_in_period if cogs_annual > 0 else 0
    ccc = dio + dso - dpo
    
    # Net Working Capital
    nwc = average_inventory + average_receivables - average_payables
    
    return {
        "dio_days": round(dio, 1),
        "dso_days": round(dso, 1),
        "dpo_days": round(dpo, 1),
        "cash_conversion_cycle_days": round(ccc, 1),
        "net_working_capital": round(nwc, 2),
        "diagnosis": (
            "Sangat Efisien (Likuiditas Tinggi)" if ccc < 30 else
            "Normal & Stabil" if ccc <= 75 else
            "Tertekan (Banyak Modal Tertahan di Piutang/Stok, Perlu Perbaikan Arus Kas)"
        )
    }


# ==============================================================================
# 4. KESEHATAN KEUANGAN KORPORAT & RISIKO KEBANGRUTAN
# ==============================================================================

def calculate_altman_z_score(
    working_capital: float,
    retained_earnings: float,
    ebit: float,
    market_or_book_value_equity: float,
    total_liabilities: float,
    sales: float,
    total_assets: float,
    is_manufacturing: bool = True
) -> Dict[str, Any]:
    """
    Menghitung Skor Altman Z untuk mendeteksi probabilitas kebangkrutan / distress finansial.
    Model Manufaktur Standar (Original Model):
    Z = 1.2*X1 + 1.4*X2 + 3.3*X3 + 0.6*X4 + 0.999*X5
    """
    if total_assets <= 0:
        raise ValueError("Total Assets harus lebih besar dari 0.")
    
    x1 = working_capital / total_assets
    x2 = retained_earnings / total_assets
    x3 = ebit / total_assets
    x4 = market_or_book_value_equity / total_liabilities if total_liabilities > 0 else 10.0
    x5 = sales / total_assets
    
    if is_manufacturing:
        z_score = (1.2 * x1) + (1.4 * x2) + (3.3 * x3) + (0.6 * x4) + (0.999 * x5)
        if z_score > 2.99:
            zone = "SAFE ZONE (Perusahaan Sangat Sehat, Risiko Kebangkrutan Sangat Rendah)"
        elif z_score >= 1.81:
            zone = "GREY ZONE (Zona Waspada, Perlu Pemantauan Modal & Restrukturisasi)"
        else:
            zone = "DISTRESS ZONE (Zona Bahaya, Probabilitas Kebangkrutan Tinggi)"
    else:
        # Non-Manufacturing / Emerging Market Model
        z_score = (6.56 * x1) + (3.26 * x2) + (6.72 * x3) + (1.05 * x4)
        if z_score > 2.60:
            zone = "SAFE ZONE (Perusahaan Sehat)"
        elif z_score >= 1.10:
            zone = "GREY ZONE (Zona Waspada)"
        else:
            zone = "DISTRESS ZONE (Zona Bahaya)"
            
    return {
        "z_score": round(z_score, 2),
        "zone_classification": zone,
        "ratios": {
            "x1_working_capital_to_assets": round(x1, 4),
            "x2_retained_earnings_to_assets": round(x2, 4),
            "x3_ebit_to_assets": round(x3, 4),
            "x4_equity_to_liabilities": round(x4, 4),
            "x5_asset_turnover": round(x5, 4)
        }
    }


def calculate_dupont_analysis(
    net_income: float,
    revenue: float,
    total_assets: float,
    total_equity: float
) -> Dict[str, Any]:
    """
    Analisis DuPont 3-Komponen untuk mengurai pendorong utama Return on Equity (ROE).
    ROE = Net Profit Margin * Asset Turnover * Financial Leverage Multiplier
    """
    if revenue <= 0 or total_assets <= 0 or total_equity <= 0:
        raise ValueError("Revenue, Total Assets, dan Total Equity harus lebih besar dari 0.")
    
    net_profit_margin = net_income / revenue
    asset_turnover = revenue / total_assets
    equity_multiplier = total_assets / total_equity
    
    roe = net_profit_margin * asset_turnover * equity_multiplier
    roa = net_profit_margin * asset_turnover
    
    return {
        "roe_pct": round(roe * 100, 2),
        "roa_pct": round(roa * 100, 2),
        "net_profit_margin_pct": round(net_profit_margin * 100, 2),
        "asset_turnover_ratio": round(asset_turnover, 2),
        "equity_multiplier_leverage": round(equity_multiplier, 2),
        "executive_summary": f"ROE tercatat {round(roe * 100, 2)}%, didorong oleh Margin Bersih {round(net_profit_margin * 100, 1)}%, Perputaran Aset {round(asset_turnover, 2)}x, dan Leverage Keuangan {round(equity_multiplier, 2)}x."
    }


# ==============================================================================
# 5. CAPITAL BUDGETING & KELAYAKAN INVESTASI (NPV, IRR, PAYBACK)
# ==============================================================================

def evaluate_investment_project(
    initial_outlay: float,
    annual_cash_flows: List[float],
    discount_rate: float
) -> Dict[str, Any]:
    """
    Evaluasi Proyek Capex / Investasi Baru (NPV, IRR, Payback Period, PI).
    
    :param initial_outlay: Nilai investasi awal (positif, misal Rp 200.000.000).
    :param annual_cash_flows: Arus kas masuk tahunan [Thn 1, Thn 2, ...].
    :param discount_rate: Hurdle rate / WACC (misal 0.12 untuk 12%).
    """
    import numpy_financial as npf
    full_cash_flows = [-initial_outlay] + annual_cash_flows
    
    # Net Present Value (NPV)
    pv_inflows = [cf / ((1.0 + discount_rate) ** t) for t, cf in enumerate(annual_cash_flows, start=1)]
    npv = sum(pv_inflows) - initial_outlay
    
    # Internal Rate of Return (IRR)
    try:
        irr_val = npf.irr(full_cash_flows)
    except Exception:
        irr_val = None
            
    # Profitability Index (PI)
    pi = sum(pv_inflows) / initial_outlay if initial_outlay > 0 else 0
    
    # Payback Period (Non-discounted)
    cumulative_cf = 0.0
    payback_years = None
    for t, cf in enumerate(annual_cash_flows, start=1):
        cumulative_cf += cf
        if cumulative_cf >= initial_outlay:
            fraction = (initial_outlay - (cumulative_cf - cf)) / cf
            payback_years = (t - 1) + fraction
            break
            
    recommendation = (
        "DISETUJUI / FEASIBLE (NPV Positif & IRR di atas Hurdle Rate)"
        if npv > 0 and (irr_val is not None and irr_val > discount_rate)
        else "DITOLAK / TIDAK LAYAK (NPV Negatif atau Return di bawah Hurdle Rate)"
    )
    
    return {
        "initial_investment": round(initial_outlay, 2),
        "npv": round(npv, 2),
        "irr_pct": round(irr_val * 100, 2) if irr_val is not None else None,
        "profitability_index": round(pi, 2),
        "payback_period_years": round(payback_years, 2) if payback_years is not None else "Lebih dari horizon proyek",
        "hurdle_rate_pct": round(discount_rate * 100, 2),
        "recommendation": recommendation
    }


# ==============================================================================
# 6. EXCEL EXECUTIVE REPORT GENERATOR DENGAN SANITASI ISO 9001
# ==============================================================================

def export_executive_financial_dashboard(
    filepath: str,
    project_title: str,
    valuation_data: Dict[str, Any],
    unit_econ_data: Dict[str, Any],
    dupont_data: Dict[str, Any]
) -> str:
    """
    Menghasilkan Master Dashboard Excel Eksekutif berstandar ISO 9001:2015
    dengan sanitasi metadata resmi PT Kediri Chemical Abadi.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    ws.views.sheetView[0].showGridLines = True
    
    # Styling Palette Korporat Resmi
    navy_fill = PatternFill(start_color="0E2A47", end_color="0E2A47", fill_type="solid")
    accent_blue_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=11, bold=True, color="000000")
    font_bold_black = Font(name="Calibri", size=10, bold=True, color="000000")
    font_regular_black = Font(name="Calibri", size=10, regular=True, color="000000")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # 1. Header Banner Korporat
    ws.merge_cells("A1:F2")
    ws["A1"] = f"PT KEDIRI CHEMICAL ABADI — EXECUTIVE DASHBOARD & STRATEGIC ADVISORY\n{project_title.upper()}"
    ws["A1"].font = font_title
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 2. Metadata Pengendali Dokumen ISO 9001
    ws["A3"] = "Penanggung Jawab: Yerikho Arfensias Effendi"
    ws["D3"] = "Direktur Utama: Yan Effendi"
    ws["A4"] = "Klasifikasi: Rahasia / Keputusan Direksi"
    ws["D4"] = "Standar: ISO 9001:2015 Terkendali"
    for r in range(3, 5):
        for c in range(1, 7):
            ws.cell(row=r, column=c).font = Font(name="Calibri", size=9, italic=True, color="333333")
            
    # Helper Penulisan Metrik
    curr_row = 6
    
    def write_section(title: str, items: List[Tuple[str, Any]]):
        nonlocal curr_row
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
        c = ws.cell(row=curr_row, column=1, value=f"▶ {title.upper()}")
        c.font = font_section
        c.fill = accent_blue_fill
        curr_row += 1
        
        for k, v in items:
            ws.cell(row=curr_row, column=1, value=k).font = font_regular_black
            val_cell = ws.cell(row=curr_row, column=3, value=v)
            val_cell.font = font_bold_black
            val_cell.alignment = Alignment(horizontal="right")
            for col in range(1, 7):
                ws.cell(row=curr_row, column=col).border = thin_border
            curr_row += 1
        curr_row += 1
        
    # Tulis Seksi Valuasi
    write_section("1. Valuasi Korporat (Discounted Cash Flow)", [
        ("Enterprise Value (EV)", f"Rp {valuation_data.get('enterprise_value', 0):,.2f}"),
        ("Net Debt", f"Rp {valuation_data.get('net_debt', 0):,.2f}"),
        ("Equity Value", f"Rp {valuation_data.get('equity_value', 0):,.2f}"),
        ("Terminal Value % of EV", f"{valuation_data.get('pv_terminal_value_pct', 0)}%"),
    ])
    
    # Tulis Seksi Unit Economics
    write_section("2. Unit Economics & Efisiensi Akuisisi", [
        ("LTV / CAC Ratio", f"{unit_econ_data.get('ltv_cac_ratio', 0)}x"),
        ("CAC Payback Period", f"{unit_econ_data.get('cac_payback_months', 0)} Bulan"),
        ("Customer Lifetime Value (LTV)", f"Rp {unit_econ_data.get('customer_lifetime_value_ltv', 0):,.2f}"),
        ("Status Kesehatan Bisnis", str(unit_econ_data.get('health_status', 'N/A')))
    ])
    
    # Tulis Seksi DuPont ROE
    write_section("3. DuPont 3-Way ROE Decomposition", [
        ("Return on Equity (ROE)", f"{dupont_data.get('roe_pct', 0)}%"),
        ("Net Profit Margin", f"{dupont_data.get('net_profit_margin_pct', 0)}%"),
        ("Asset Turnover", f"{dupont_data.get('asset_turnover_ratio', 0)}x"),
        ("Financial Leverage Multiplier", f"{dupont_data.get('equity_multiplier_leverage', 0)}x")
    ])
    
    # Adjust Column Widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Sanitasi Metadata Excel
    wb.properties.creator = "Yerikho Arfensias Effendi"
    wb.properties.lastModifiedBy = "Yerikho Arfensias Effendi"
    wb.properties.title = project_title
    wb.properties.subject = "Laporan Konsultasi Bisnis Strategis & Keputusan CEO"
    wb.properties.company = "PT Kediri Chemical Abadi"
    wb.properties.category = "Laporan Eksekutif Direksi"
    
    wb.save(filepath)
    return filepath


if __name__ == "__main__":
    print("CEO Consulting Toolkit Initialized.")
